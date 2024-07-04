import os
import argparse

import pandas
from tqdm import tqdm

import conllu_parser
import statifier
import proximity
import parse_conllu

import openpyxl
import openpyxl.styles
import math

UDDIR = "ud-treebanks-v2.14"

parser = argparse.ArgumentParser()
parser.add_argument("-fi", "--filename", help=".conllu file containing a UD-Treebank. Use all to iterate on all banks.")
parser.add_argument("-v", "--verbose", required=False)
parser.add_argument("-p", "--property", help="UD-RELDEP which we're studying, e.g. obl:tmod", required=False)
parser.add_argument("-f", "--feature_type", help="Morphosyntactic category we want to study.", required=False)
parser.add_argument("-val", "--value", help="Value of the studied feature", required=False)
parser.add_argument("-o", "--out", help="File in which to store the result", required=False, default="Sentence_Graphs")
parser.add_argument(
    "-i", "--index",
    help=f"Number of the sentences to study. Must be smaller than MAX_STUDIED_SENTENCES={conllu_parser.MAX_STUDIED_SENTENCES}",
    required=False
)
args = parser.parse_args()


def empacker(filename, reldep, gf):
    files = filename.split("/")
    pwd = f"{UDDIR}/{files[0]}"
    out = pwd + f"/{files[1][:-7]}/"
    try:
        os.mkdir(out)
    except FileExistsError:
        pass

    for c in ["/RelDep_Matches", "/Grammar_Matches"]:
        try:
            os.mkdir(out + c)
        except FileExistsError:
            pass

    grammar_feature_for_reldep, reldep_for_grammar_feature = parse_conllu.vectorize(
        f"{UDDIR}/{filename}", grammar_feature, reldep, out=None
        )

    with open(f"{out}/Grammar_Matches/{gf[0]}_matching_{reldep}", 'w') as f:
        f.write(grammar_feature_for_reldep)

    with open(f"{out}/RelDep_Matches/RelDep_matching_{gf[0]}={gf[1]}.txt", "w") as f:
        f.write(reldep_for_grammar_feature)


def empacking(filename, reldep, gf, index, verbose):
    filename = filename.split("/")

    pwd = f"{UDDIR}/{filename[0]}"
    out = pwd + f"/{filename[1][:-7]}/"
    try:
        os.mkdir(out)
    except FileExistsError:
        pass

    for c in ["/Graph_Sources", "/Graphs", "/UD_RelDep", "/Features", "/RelDep_Matches", "/Grammar_Matches"]:
        try:
            os.mkdir(out + c)
        except FileExistsError:
            pass

    number_of_sentences, cpt = conllu_parser.tree_ifier(
        pwd + "/" + filename[1], ud_reldep=reldep, grammar_feature=gf, out=out
        )
    if index:
        conllu_parser.show_graph(index, "deep/" + filename[0] + "/" + args.out)
    right_type_dict, other_type_dict = statifier.frequency_on_grammar_feature_checking_reldep(
        out + f"/UD_RelDep/{args.property}.txt",
        args.feature_type
    )
    property_matches = (sum(map(lambda l: l[1], right_type_dict.items())))
    reldep_dict = statifier.frequency_on_reldep_checking_grammar_feature(
        f"{out}/Features/{gf[0]}={gf[1]}.txt"
    )

    reldep_for_grammar_feature = ""
    grammar_feature_for_reldep = f"We have studied {number_of_sentences} sentences and failed on {cpt} in treebank `{filename}`.\n"
    grammar_feature_for_reldep += f"Of those, {property_matches} words match `{reldep}`.\n"
    if property_matches != 0:
        right_type_dict = statifier.format_case_stats(right_type_dict)
        grammar_feature_for_reldep += (f"We get the following distribution of values for `{gf[0]}` "
                                       f"matching `{reldep}`:\n")
        for v in right_type_dict:
            grammar_feature_for_reldep += f"\t{v}: {right_type_dict[v]}\n"

        if right_type_dict.get("Other_Word_Type", 0) != 0:
            grammar_feature_for_reldep += ("The Other Words have the following distribution of grammatical "
                                           "features:\n")
            other_type_dict = statifier.format_property_stats(other_type_dict)
            for v in other_type_dict:
                grammar_feature_for_reldep += f"Feature {v}:\n"
                for a in other_type_dict[v]:
                    grammar_feature_for_reldep += f"\t{a}: {other_type_dict[v][a]}\n"
        with open(f"{out}/Grammar_Matches/{gf[0]}_matching_{reldep}", 'w') as f:
            f.write(grammar_feature_for_reldep)
        reldep_for_grammar_feature += (f"We have studied {number_of_sentences} sentences and failed on {cpt} in "
                                       f"treebank `{filename}`.\n We get the following distribution "
                                       f"of RelDep for words matching `{gf[0]}={gf[1]}`:\n")
        reldep_dict = statifier.format_case_stats(reldep_dict)
        for v in reldep_dict:
            reldep_for_grammar_feature += f"RelDep {v}: {reldep_dict[v]}\n"
        with open(f"{out}/RelDep_Matches/RelDep_matching_{gf[0]}={gf[1]}.txt", "w") as f:
            f.write(reldep_for_grammar_feature)
    if verbose:
        print(grammar_feature_for_reldep)
        if property_matches != 0:
            print(reldep_for_grammar_feature)


def fmc(c):
    """
    Function used to format a value from an Excel cell.
    :param c: String collected from an Excel table
    :return: Integer representing the value
    """
    if c is None:
        return 0
    return int(c)


def pre_process():
    if args.filename != "all":
        empacker(args.filename, args.property, grammar_feature)
    else:
        treebanks = os.listdir(UDDIR)
        reldep_dict = {}
        number_of_studied_sentences = 0
        cpt = 0
        for treebank in (pbar := tqdm(treebanks, colour='#7d1dd3')):
            pbar.set_description(f"Processing {treebank}")
            content = os.listdir(f"{UDDIR}/{treebank}")
            for c in list(filter(lambda t: t[-7:] == ".conllu", content)):
                try:
                    with open(
                            f"{UDDIR}/{treebank}/{c[:-7]}/RelDep_Matches/RelDep_matching_{grammar_feature[0]}={grammar_feature[1]}.txt",
                            'r'
                    ) as f:
                        reldeps = f.readlines()

                    for i in range(len(reldeps)):
                        reldeps[i] = reldeps[i].split(" ")

                    number_of_studied_sentences += int(reldeps[0][3])
                    cpt += int(reldeps[0][8])

                    for i in range(2, len(reldeps)):
                        rel = reldeps[i][1][:-1]
                        reldep_dict[rel] = reldep_dict.get(rel, 0) + int(reldeps[i][2])

                except FileNotFoundError:
                    empacker(f"{treebank}/{c}", args.property, grammar_feature)
                    try:
                        with open(
                                f"{UDDIR}/{treebank}/{c[:-7]}/RelDep_Matches/RelDep_matching_{grammar_feature[0]}={grammar_feature[1]}.txt",
                                'r'
                        ) as f:
                            reldeps = f.readlines()
                        for i in range(len(reldeps)):
                            reldeps[i] = reldeps[i].split(" ")
                        number_of_studied_sentences += int(reldeps[0][3])
                        cpt += int(reldeps[0][8])
                        for i in range(2, len(reldeps)):
                            rel = reldeps[i][1][:-1]
                            reldep_dict[rel] = reldep_dict.get(rel, 0) + int(reldeps[i][2])
                    except FileNotFoundError:
                        pass

        pbar.set_description("Treebank processing done")
        reldep_for_grammar_feature = (f"We have studied {number_of_studied_sentences} sentences and failed on {cpt} in "
                                      f"all treebanks.\nWe get the following distribution "
                                      f"of RelDep for words matching `{grammar_feature[0]}={grammar_feature[1]}`:\n")
        reldep_dict = statifier.format_case_stats(reldep_dict)
        for v in reldep_dict:
            reldep_for_grammar_feature += f"RelDep {v}: {reldep_dict[v]}\n"
        with open(f"RelDep_Matches/RelDep_matching_{grammar_feature[0]}={grammar_feature[1]}.txt", "w") as f:
            f.write(reldep_for_grammar_feature)


def from_reldep_to_table(rel_dep_matching_grammar_feature, wb):
    if rel_dep_matching_grammar_feature in wb.sheetnames:
        return
    print("Tabulating ", rel_dep_matching_grammar_feature)
    wb.create_sheet(title=rel_dep_matching_grammar_feature)
    ws = wb[rel_dep_matching_grammar_feature]
    rel_dep_dict = {}
    max_column = 0
    max_row = 4
    ws.cell(1, 1).value = "Treebank"
    ws.cell(2, 1).value = "Number of Studied Sentences"
    ws.cell(3, 1).value = "Number of Failed Sentences"

    sum_column = ["Sum:"]
    for c in (pbar := tqdm(
            list(
                filter(
                    lambda t: t[1][-7:] == ".conllu",
                    [(treebanks, treebank) for treebanks in os.listdir(UDDIR) for treebank in
                     os.listdir(f"{UDDIR}/{treebanks}")]
                )
            )
            , colour="#7d1dd3"
    )):
        treebanks, treebank = c
        treebank = treebank[:-7]
        pbar.set_description(f"Tabulating {treebanks}/{treebank}")

        try:
            with open(
                    f"{UDDIR}/{treebanks}/{treebank}/RelDep_Matches/{rel_dep_matching_grammar_feature}.txt"
            ) as f:
                results = f.readlines()
            results[0] = results[0].split(" ")
            number_of_sentences = results[0][3]
            cpt_sentences = results[0][8]
            title = treebank
            ws.cell(1, max_column + 2).value = title
            ws.cell(2, max_column + 2).value = number_of_sentences
            ws.cell(3, max_column + 2).value = cpt_sentences
            total_values = 0
            for line in results[2:]:
                res = line.split(" ")
                rel_dep = (res[1][:-1])
                number = res[2]
                # print(number)
                if rel_dep not in rel_dep_dict:
                    ws.cell(max_row, 1).value = rel_dep
                    ws.cell(max_row, max_column + 2).value = number
                    rel_dep_dict[rel_dep] = max_row
                    max_row += 1
                else:
                    ws.cell(rel_dep_dict[rel_dep], max_column + 2).value = number
                total_values += number
            max_column += 1
        except FileNotFoundError:
            title = treebank
            ws.cell(1, max_column + 2).value = title
            max_column += 1
            total_values = 0
        sum_column.append(total_values)

    pbar.set_description("Tabulating Treebanks Done")
    ws.append(sum_column)


def from_reldep_to_csv(rel_dep_matching_grammar_feature):
    try:
        open(f"RelDep_Matches/{rel_dep_matching_grammar_feature}.csv", "r")
    except FileNotFoundError:
        print("Tabulating ", rel_dep_matching_grammar_feature)
        value_dicts = []
        fieldnames = set()

        for c in (pbar := tqdm(
                list(
                    filter(
                        lambda t: t[1][-7:] == ".conllu",
                        [(treebanks, treebank) for treebanks in os.listdir(UDDIR) for treebank in
                         os.listdir(f"{UDDIR}/{treebanks}")]
                    )
                )
                , colour="#7d1dd3"
        )):
            vec_coordinates = {}
            treebanks, treebank = c
            treebank = treebank[:-7]
            pbar.set_description(f"Tabulating {treebanks}/{treebank}")
            vec_coordinates["Treebank"] = treebank
            try:
                with open(
                        f"{UDDIR}/{treebanks}/{treebank}/RelDep_Matches/{rel_dep_matching_grammar_feature}.txt"
                ) as f:
                    results = f.readlines()
                results[0] = results[0].split(" ")

                vec_coordinates["Number of Sentences"] = results[0][3]
                vec_coordinates["Failures"] = results[0][8]
                total_values = 0
                for line in results[2:]:
                    res = line.split(" ")
                    reldep = res[1][:-1]
                    number = res[2]
                    vec_coordinates[reldep] = int(number)
                    total_values += int(number)
                    fieldnames.add(reldep)
                vec_coordinates["Total"] = total_values
            except FileNotFoundError:
                pass
            value_dicts.append(vec_coordinates)
        pbar.set_description("Tabulating Treebanks Done")
        pbar.close()
        pandas.DataFrame(value_dicts).to_csv(f"RelDep_Matches/{rel_dep_matching_grammar_feature}.csv", index=False)


def re_process_xl():
    wb = openpyxl.load_workbook("RelDep_Matches_old.xlsx")
    for rel_dep_matching_grammar_feature in os.listdir("RelDep_Matches"):
        from_reldep_to_table(rel_dep_matching_grammar_feature[:-4], wb)
    wb.save("RelDep_Matches.xlsx")


def re_process_csv():
    for rel_dep_matching_grammar_feature in os.listdir("RelDep_Matches"):
        from_reldep_to_csv(rel_dep_matching_grammar_feature[:-4])


if __name__ == "__main__":
    print(f"Processing: RelDep={args.property}, {args.feature_type}={args.value}")
    grammar_feature = (args.feature_type, args.value)
    if args.filename:
        pre_process()
    re_process_csv()
