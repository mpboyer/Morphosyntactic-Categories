import math
import joblib
import openpyxl
import openpyxl.styles
import pandas
import scipy.linalg
from tqdm import tqdm, trange
import numpy as np
import numpy.linalg as npl
from sympy import Matrix
import os
import contextlib
import argparse

from linalg import project, angle, distance, is_in_cone, manhattan_normalizer

UDDIR = "ud-treebanks-v2.14"

parser = argparse.ArgumentParser()
parser.add_argument("-c", "--case-space")
parser.add_argument("-t", "--treebank-case")
args = parser.parse_args()


@contextlib.contextmanager
def tqdm_joblib(tqdm_object):
    """Context manager to patch joblib to report into tqdm progress bar given as argument"""

    class TqdmBatchCompletionCallback(joblib.parallel.BatchCompletionCallBack):
        def __call__(self, *args, **kwargs):
            tqdm_object.update(n=self.batch_size)
            return super().__call__(*args, **kwargs)

    old_batch_callback = joblib.parallel.BatchCompletionCallBack
    joblib.parallel.BatchCompletionCallBack = TqdmBatchCompletionCallback
    try:
        yield tqdm_object
    finally:
        joblib.parallel.BatchCompletionCallBack = old_batch_callback
        tqdm_object.close()


def get_all_banks():
    abank = []
    treebanks = os.listdir(UDDIR)
    for treebank in treebanks:
        content = os.listdir(f"{UDDIR}/{treebank}")
        for c in list(filter(lambda t: t[-7:] == ".conllu", content)):
            abank.append(c[:-7])
    return abank


def is_name_in_sheets(sheetname, filename):
    workbook = openpyxl.load_workbook(filename)
    if sheetname in workbook.sheetnames:
        return True
    workbook.close()
    return False


def tabulize_xl(grammar_feature):
    if is_name_in_sheets(f"Proximity_Stats_for_{grammar_feature[0]}={grammar_feature[1]}", "Proximity.xlsx"):
        return
    print(f"Tabulizing for {grammar_feature[0]}={grammar_feature[1]}")
    database = pandas.read_excel(
        "RelDep_Matches.xlsx",
        sheet_name=f"RelDep_matching_{grammar_feature[0]}={grammar_feature[1]}"
    )
    database.fillna(0, inplace=True)
    # max_row = len(database["Treebank"])
    for column in (pbar := tqdm(database, colour='#7d1dd3', total=574)):
        if column == "Treebank":
            pass
        else:
            pbar.set_description(f"Normalizing: {column}")
            c = [database[column][i] for i in range(2, len(database[column]) - 1)]
            c = manhattan_normalizer(c)
            for i in range(len(c)):
                database.loc[i + 2, column] = c[i]
    workbook = openpyxl.load_workbook("Proximity_old.xlsx")
    worksheet = workbook.create_sheet(f"Proximity_Stats_for_{grammar_feature[0]}={grammar_feature[1]}")
    worksheet.cell(1, 1).value = "Treebank"
    treetab = {}
    mat = database[database.columns[1:]].to_numpy()[2:-1]
    for i in range(len(mat[0])):
        euclidean_norm = npl.norm(mat[:, i])
        if euclidean_norm != 0:
            mat[:, i] /= euclidean_norm
    dot_mat = np.matmul(np.transpose(mat), mat)
    i = 2
    for (vec1, vec2) in [(vec1, vec2) for vec1 in database for vec2 in database]:
        if vec1 == "Treebank":
            pass
        else:
            if vec1 not in treetab:
                worksheet.cell(1, i).value = vec1
                worksheet.cell(i, 1).value = vec1
                treetab[vec1] = i
                i += 1
            if vec2 == "Treebank":
                pass
            else:
                if vec2 not in treetab:
                    worksheet.cell(1, i).value = vec2
                    worksheet.cell(i, 1).value = vec2
                    treetab[vec2] = i
                    i += 1
                row = treetab[vec1]
                column = treetab[vec2]
                euclidean_norm = dot_mat[row - 2, column - 2]
                worksheet.cell(row, column).value = euclidean_norm
                worksheet.cell(row, column).fill = openpyxl.styles.fills.PatternFill(
                    patternType='solid',
                    fgColor=openpyxl.styles.colors.Color(
                        indexed=math.floor(euclidean_norm * 5 + 2)
                    )
                )
    workbook.save("Proximity.xlsx")


def tabulize_pair_xl(gf1, gf2, result_workbook):
    if f"Proximity_{gf1[0]}={gf1[1]}_{gf2[0]}={gf2[1]}" in result_workbook.sheetnames:
        return
    workbook = openpyxl.load_workbook("RelDep_Matches.xlsx")
    reldep_table = {}
    current_row = 0
    worksheet = workbook[f"RelDep_matching_{gf1[0]}={gf1[1]}"]
    max_col1 = worksheet.max_column + 1
    max_row1 = worksheet.max_row
    all_vec = [{} for _ in range(2, max_col1)]
    for working_column in range(2, max_col1):
        all_vec[working_column - 2]["Name"] = f"{worksheet.cell(1, working_column).value}_{gf1[0]}={gf1[1]}"
        all_vec[working_column - 2]["Sum"] = worksheet.cell(max_row1, working_column).value
    pbar = tqdm(total=(max_row1 - 4) * (max_col1 - 2), colour="#7d1dd3", desc=f"Reading {gf1}")
    for working_row in range(4, max_row1):
        reldep = worksheet.cell(working_row, 1).value
        if reldep not in reldep_table:
            reldep_table[reldep] = current_row
            current_row += 1
        for working_column in range(2, max_col1):
            if all_vec[working_column - 2]["Sum"]:
                dot = int(worksheet.cell(working_row, working_column).value) if worksheet.cell(
                    working_row, working_column
                ).value else 0
                all_vec[working_column - 2][reldep] = dot / all_vec[working_column - 2]["Sum"]
            pbar.update(1)
    pbar.close()
    current_column = len(all_vec)

    worksheet = workbook[f"RelDep_matching_{gf2[0]}={gf2[1]}"]
    max_col2 = worksheet.max_column + 1
    max_row2 = worksheet.max_row
    all_vec = all_vec + [{} for _ in range(2, max_col2)]
    for working_column in range(2, max_col2):
        all_vec[current_column + working_column - 2][
            "Name"] = f"{worksheet.cell(1, working_column).value}_{gf2[0]}={gf2[1]}"
    pbar = tqdm(total=(max_row2 - 4) * (max_col2 - 2), colour="#7d1dd3", desc=f"Reading {gf2}")
    for working_row in range(4, max_row2):
        reldep = worksheet.cell(working_row, 1).value
        if reldep not in reldep_table:
            reldep_table[reldep] = current_row
            current_row += 1
        for working_column in range(2, max_col2):
            if all_vec[working_column - 2]["Sum"]:
                dot = int(worksheet.cell(working_row, working_column).value) if worksheet.cell(
                    working_row, working_column
                ).value else 0
                all_vec[current_column + working_column - 2][reldep] = dot / all_vec[working_column - 2]["Sum"]
            pbar.update(1)
    pbar.close()
    current_column = len(all_vec)
    workbook.close()

    mat = np.matrix(np.array([[0. for _ in range(current_column)] for _ in range(current_row)]))
    for reldep in tqdm(reldep_table, desc="Reporting Matrix", colour="#ffe500"):
        working_row = reldep_table[reldep]
        for k in range(current_column):
            mat[working_row, k] = all_vec[k].get(reldep, 0.)

    for working_column in tqdm(range(current_column), desc="Normalizing Matrix", colour="#ffe500"):
        euclidean_norm = npl.norm(mat[:, working_column])
        if euclidean_norm != 0:
            mat[:, working_column] /= euclidean_norm
    dot_mat = np.matmul(np.transpose(mat), mat)

    worksheet = result_workbook.create_sheet(f"Proximity_{gf1[0]}={gf1[1]}_{gf2[0]}={gf2[1]}")
    worksheet.cell(1, 1).value = "Treebank+GF"
    pbar = tqdm(desc="Reporting Values in Table", total=current_column * current_column, colour="#7d1dd3")
    for working_row in range(current_column):
        worksheet.cell(1, working_row + 2).value = all_vec[working_row]["Name"]
        worksheet.cell(working_row + 2, 1).value = all_vec[working_row]["Name"]
        for working_column in range(current_column):
            dot = dot_mat[working_row, working_column]
            worksheet.cell(working_row + 2, working_column + 2).value = dot
            worksheet.cell(working_row + 2, working_column + 2).fill = openpyxl.styles.fills.PatternFill(
                patternType='solid',
                fgColor=openpyxl.styles.colors.Color(
                    indexed=math.floor(dot * 5 + 2)
                )
            )
            pbar.update(1)
    pbar.close()


def overall_basis_xl():
    all_reldeps = pandas.ExcelFile("RelDep_Matches.xlsx")
    case_sheets = [s for s in all_reldeps.sheet_names if s != "Sheet" and s[16:20] == "Case"]
    basis = {}
    for sheet in case_sheets:
        worksheet = all_reldeps.parse(sheet)
        for current_row in range(2, len(worksheet[worksheet.columns[0]]) - 1):
            basis[worksheet[worksheet.columns[0]][current_row]] = 0
    return basis


def get_matrix_xl(treebank):
    """
    :param treebank: Name of ud treebank. `treebank.conllu` must exist.
    :return: Numpy matrix containing the vector representation of the grammatical cases of the treebank. The result is in row-echelon form and is normalized for manhattan distance.
    """
    reldep_matches = pandas.ExcelFile("RelDep_Matches.xlsx")
    case_sheets = [s for s in reldep_matches.sheet_names if s != "Sheet" and s[16:20] == "Case"]
    value_dict = {}
    #    for sheet in tqdm(case_sheets, colour="#7d1dd3", desc=f"Vectorizing {treebank}"):
    for sheet in case_sheets:
        worksheet = reldep_matches.parse(sheet)
        manhattan_norm = np.nan_to_num(worksheet[treebank][len(worksheet[treebank]) - 1])
        if manhattan_norm != 0:
            for current_ in range(2, len(worksheet[treebank]) - 1):
                value = np.nan_to_num(worksheet[treebank][current_])
                if worksheet["Treebank"][current_] in value_dict:
                    value_dict[worksheet["Treebank"][current_]][sheet] = value / manhattan_norm
                else:
                    value_dict[worksheet["Treebank"][current_]] = {
                        sheet: value / manhattan_norm}
        else:
            for current_ in range(2, len(worksheet[treebank]) - 1):
                if worksheet["Treebank"][current_] in value_dict:
                    value_dict[worksheet["Treebank"][current_]][sheet] = 0.
                else:
                    value_dict[worksheet["Treebank"][current_]] = {
                        sheet: 0.}
    reldep_mat = np.matrix(np.array([[0. for _ in case_sheets] for _ in value_dict]))

    current_ = 0
    for key in value_dict:
        column = 0
        for sheet in case_sheets:
            reldep_mat[current_, column] = value_dict[key].get(sheet, 0.)
            column += 1
        current_ += 1

    return reldep_mat


def enhanced_get_matrix_xl(case_space_filename, vector_filename):
    reldep_matches = pandas.ExcelFile("RelDep_Matches.xlsx")
    case_sheets = [s for s in reldep_matches.sheet_names if s != "Sheet" and s[16:20] == "Case"]
    case_space_value_dict = {}
    vec_value_dict = {}
    #    for sheet in tqdm(case_sheets, colour="#7d1dd3", desc=f"Vectorizing {treebank}"):
    for sheet in case_sheets:
        ws = reldep_matches.parse(sheet)
        if vector_filename[1] == sheet[-len(vector_filename[1]):]:
            d = np.nan_to_num(ws[vector_filename[0]][len(ws[vector_filename[0]]) - 1])
            if d != 0:
                for current_row in range(2, len(ws[vector_filename[0]]) - 1):
                    value = np.nan_to_num(ws[vector_filename[0]][current_row])
                    vec_value_dict[ws["Treebank"][current_row]] = value / d

        d = np.nan_to_num(ws[case_space_filename][len(ws[case_space_filename]) - 1])
        if d != 0:
            for current_row in range(2, len(ws[case_space_filename]) - 1):
                value = np.nan_to_num(ws[case_space_filename][current_row])
                if ws["Treebank"][current_row] in case_space_value_dict:
                    case_space_value_dict[ws["Treebank"][current_row]][sheet] = value / d
                else:
                    case_space_value_dict[ws["Treebank"][current_row]] = {
                        sheet: value / d}
        else:
            for current_row in range(2, len(ws[case_space_filename]) - 1):
                if ws["Treebank"][current_row] in case_space_value_dict:
                    case_space_value_dict[ws["Treebank"][current_row]][sheet] = 0.
                else:
                    case_space_value_dict[ws["Treebank"][current_row]] = {
                        sheet: 0.}

    case_space = np.array([[0. for _ in case_sheets] for _ in case_space_value_dict])
    vector = np.array([0. for _ in case_space_value_dict])

    current_row = 0
    for key in case_space_value_dict:
        column = 0
        for sheet in case_sheets:
            case_space[current_row, column] = case_space_value_dict[key].get(sheet, 0.)
            column += 1
        vector[current_row] = vec_value_dict.get(key, 0.)
        current_row += 1

    return case_space, vector


def zassenhaus(m1, m2):
    """
    This function uses the Zassenhaus Algorithm to compute a basis of the sum and the
    intersection of two vector sub-spaces of a same vector space.
    :param m1: Matrix representing first vector space. Must be expressed in the same basis as m2.
    :param m2: Matrix representing second vector space. Must be expressed in the same basis as m1.
    :return: Computes a matrix representing the intersection of the aforementioned vector spaces
    """
    basis_length, col1 = m1.shape
    col2 = m2.shape[1]
    zassenhaus_matrix = np.matrix(np.array([[0. for _ in range(2 * basis_length)] for _ in range(col1 + col2)]))
    for column in range(col1):
        zassenhaus_matrix[column, :basis_length] = m1[:, column].copy().transpose()
        zassenhaus_matrix[column, basis_length:] = m1[:, column].copy().transpose()

    for column in range(col2):
        zassenhaus_matrix[col1 + column, :basis_length] = m2[:, column].copy().transpose()
        # print(zassenhaus_matrix[column])

    z_mat = Matrix(zassenhaus_matrix)
    # print(z_mat)
    z_mat = z_mat.rref()[0]
    # print(z_mat)
    zassenhaus_matrix = np.array(z_mat)
    sum_basis = []
    intersection_basis = []
    current_row = 0
    while np.count_nonzero(zassenhaus_matrix[current_row, :basis_length]) and current_row < 2 * basis_length:
        sum_basis.append(zassenhaus_matrix[current_row, :basis_length])
        current_row += 1

    while np.count_nonzero(zassenhaus_matrix[current_row, basis_length:]) and current_row < 2 * basis_length:
        intersection_basis.append(zassenhaus_matrix[current_row, basis_length:])
        current_row += 1

    return sum_basis, intersection_basis


def vector_space_proximity(treebank1, treebank2):
    mat1 = get_matrix_xl(treebank1)
    dim1 = 0
    for column in range(mat1.shape[1]):
        if np.count_nonzero(mat1[:, column]):
            dim1 += 1

    mat2 = get_matrix_xl(treebank2)
    dim2 = 0
    for column in range(mat2.shape[1]):
        if np.count_nonzero(mat2[:, column]):
            dim2 += 1

    s_b, i_b = zassenhaus(mat1, mat2)
    return treebank1, treebank2, dim1, dim2, len(s_b), len(i_b)


def get_vector_xl(vector_id):
    vector_values = pandas.ExcelFile("RelDep_Matches.xlsx").parse(f"RelDep_matching_Case={vector_id[1]}")
    vec_value_dict = {}
    manhattan_norm = np.nan_to_num(vector_values[vector_id[0]][len(vector_values[vector_id[0]]) - 1])
    if manhattan_norm != 0:
        for current_writing_row in range(2, len(vector_values[vector_id[0]]) - 1):
            value = np.nan_to_num(vector_values[vector_id[0]][current_writing_row])
            vec_value_dict[vector_values["Treebank"][current_writing_row]] = value / manhattan_norm
    return vec_value_dict


def compute_distances_xl():
    for i in range(len(gf)):
        case1 = gf[i]
        try:
            wb = openpyxl.load_workbook(f"DuoProximity/Distance_{case1}.xlsx")
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        for j in range(0, len(gf)):
            case2 = gf[j]
            if f"Distance_{case2}" not in wb.sheetnames:
                print("Distancing", case1, case2)
                rows = {}
                row = 2
                cols = {}
                col = 2
                ws = wb.create_sheet(f"Distance_{case2}")
                reldep_matches_database = pandas.ExcelFile("RelDep_Matches.xlsx")
                case_tabulating_sheets = [s for s in reldep_matches_database.sheet_names if
                                          s != "Sheet" and s[16:20] == "Case"]
                for case in case_tabulating_sheets:
                    if case[-3:] == case1:
                        case1_sheet = reldep_matches_database.parse(case)
                    if case[-3:] == case2:
                        case2_sheet = reldep_matches_database.parse(case)

                def treebank_distance(treebank1, treebank2):
                    v1 = {}
                    manhattan_norm = np.nan_to_num(case1_sheet[treebank1][len(case1_sheet[treebank1]) - 1])
                    if manhattan_norm != 0:
                        for current_writing_row in range(2, len(case1_sheet[treebank1]) - 1):
                            value = np.nan_to_num(case1_sheet[treebank1][current_writing_row])
                            v1[case1_sheet["Treebank"][current_writing_row]] = value / manhattan_norm
                    v2 = {}
                    manhattan_norm = np.nan_to_num(case2_sheet[treebank2][len(case2_sheet[treebank2]) - 1])
                    if manhattan_norm != 0:
                        for current_writing_row in range(2, len(case2_sheet[treebank2]) - 1):
                            value = np.nan_to_num(case2_sheet[treebank2][current_writing_row])
                            v2[case2_sheet["Treebank"][current_writing_row]] = value / manhattan_norm
                    return treebank1 + f"_Case={case1}", treebank2 + f"_Case={case2}", distance(v1, v2)

                distances = list(
                    joblib.Parallel(n_jobs=8, verbose=100)(
                        joblib.delayed(treebank_distance)(t1, t2) for t1 in case1_sheet.columns[1:] for t2 in
                        case2_sheet.columns[2:]
                    )
                )
                for distance_proximity in distances:
                    if distance_proximity[1] in rows:
                        cur_row = rows[distance_proximity[1]]
                    else:
                        cur_row = row
                        rows[distance_proximity[1]] = row
                        ws.cell(row, 1).value = distance_proximity[1]
                        row += 1
                    if distance_proximity[0] in cols:
                        cur_col = cols[distance_proximity[0]]
                    else:
                        cur_col = col
                        cols[distance_proximity[0]] = col
                        ws.cell(1, col).value = distance_proximity[0]
                        col += 1
                    ws.cell(cur_row, cur_col).value = distance_proximity[2]
        wb.save(f"DuoProximity/Distance_{case1}.xlsx")

    return


def compute_vector_case_space_angles_xl():
    angles = []
    all_banks = get_all_banks()
    for corpus in tqdm(all_banks, colour="#7d1dd3", leave=True, desc="Computing Angles"):
        reldep_matches_database = pandas.ExcelFile("RelDep_Matches.xlsx")
        case_tabulating_sheets = [s for s in reldep_matches_database.sheet_names if s != "Sheet" and s[16:20] == "Case"]
        mat_value_dict = {}
        for case in case_tabulating_sheets:
            ws = reldep_matches_database.parse(case)
            d = np.nan_to_num(ws[corpus][len(ws[corpus]) - 1])
            if d != 0:
                for row in range(2, len(ws[corpus]) - 1):
                    val = np.nan_to_num(ws[corpus][row])
                    if ws["Treebank"][row] in mat_value_dict:
                        mat_value_dict[ws["Treebank"][row]][case] = val / d
                    else:
                        mat_value_dict[ws["Treebank"][row]] = {
                            case: val / d}
            else:
                for row in range(2, len(ws[corpus]) - 1):
                    if ws["Treebank"][row] in mat_value_dict:
                        mat_value_dict[ws["Treebank"][row]][case] = 0.
                    else:
                        mat_value_dict[ws["Treebank"][row]] = {
                            case: 0.}

        def get_angle(case_space_filename, case_space_value_dict, vector_filename):
            vec_value_dict = get_vector_xl(vector_filename)
            case_space = np.array([[0. for _ in case_tabulating_sheets] for _ in case_space_value_dict])
            vector = np.array([0. for _ in case_space_value_dict])

            current_writing_row = 0
            for key in case_space_value_dict:
                column = 0
                for sheet in case_tabulating_sheets:
                    case_space[current_writing_row, column] = case_space_value_dict[key].get(sheet, 0.)
                    column += 1
                vector[current_writing_row] = vec_value_dict.get(key, 0.)
                current_writing_row += 1

            dimension = scipy.linalg.orth(case_space)
            if dimension.shape[1]:
                a = angle(vector, dimension)
            else:
                a = np.nan
            return case_space_filename, vector_filename[0] + f"_Case={vector_filename[1]}", a

        with tqdm_joblib(
                tqdm(
                    desc=f"Angles for {corpus}", leave=False, colour="#ffe500", position=1,
                    total=len(all_banks) * len(gf)
                )
        ) as progress_bar:
            angles += list(
                joblib.Parallel(n_jobs=8, verbose=0)(
                    joblib.delayed(get_angle)(corpus, mat_value_dict, [all_banks[j], c]) for j in range(len(all_banks))
                    for c in gf
                )
            )
    wb = openpyxl.load_workbook("Proximity.xlsx")
    rows = {}
    row = 2
    cols = {}
    col = 2
    ws = wb.create_sheet("Vector_Angle_Proximity")
    for angle_proximity in angles:
        if angle_proximity[1] in rows:
            cur_row = rows[angle_proximity[1]]
        else:
            cur_row = row
            rows[angle_proximity[1]] = row
            ws.cell(row, 1).value = angle_proximity[1]
            row += 1

        if angle_proximity[0] in cols:
            cur_col = cols[angle_proximity[0]]
        else:
            cur_col = col
            cols[angle_proximity[0]] = col
            ws.cell(1, col).value = angle_proximity[0]
            col += 1
        ws.cell(cur_row, cur_col).value = angle_proximity[2]
    wb.save("Proximity.xlsx")


def overall_basis_csv():
    basis = set()
    for csv in filter(lambda t: t[-4:] == ".csv", os.listdir("RelDep_Matches")):
        with open(f"RelDep_Matches/{csv}", "r") as csv_file:
            attributes = csv_file[0].rstrip().split(",")[4:]
        basis |= set(attributes)
    return basis


def get_vector_csv(treebank, case):
    with open(f"RelDep_Matches/RelDep_matching_Case={case}.csv", "r") as csv_file:
        attributes = csv_file[0].split(",")[4:]
        for tree in filter(lambda t: t[0] == treebank, csv_file):
            coordinates = next(csv_file).rstrip().split(",")[4:]
    return dict(zip(attributes, coordinates))


def get_matrix_csv(treebank):
    basis = set()
    case_space = []
    for csv in filter(lambda t: t[-4:] == ".csv", os.listdir("RelDep_Matches")):
        with open(f"RelDep_Matches/{csv}", "r") as csv_file:
            attributes = next(csv_file).rstrip().split(",")[4:]
            for tree in filter(lambda t: t[0] == treebank, csv_file):
                coordinates = tree.rstrip().split(",")[4:]
        case_space.append(dict(zip(attributes, coordinates)))
        basis |= set(attributes)

    matrix = np.array([[0. for _ in case_space] for _ in basis])
    for row, b in enumerate(basis):
        for column, vector in enumerate(case_space):
            matrix[row, column] = vector.get(b, 0.)

    return matrix


def void_to_zero(s):
    if s == '':
        return 0.
    return float(s)


def enhanced_get_matrix_csv(treebank, treebank_case):
    basis = set()
    case_space_vectors = []
    case_dict = {}
    for csv in sorted(filter(lambda t: t[-4:] == ".csv" and t[16:20] == "Case", os.listdir("RelDep_Matches"))):
        case_coordinates = {
            "Total": 0.}
        with open(f"RelDep_Matches/{csv}", "r") as csv_file:
            attributes = next(csv_file).rstrip().split(",")[4:]

            for tree in csv_file:
                parsed_tree = tree.rstrip().split(",")
                if csv[-7:-4] == treebank_case[1]:
                    if parsed_tree[0] == treebank_case[0]:
                        case_dict = dict(zip(attributes, map(void_to_zero, parsed_tree[4:])))
                        case_dict["Total"] = void_to_zero(parsed_tree[3])

                if parsed_tree[0] == treebank:
                    case_coordinates = dict(zip(attributes, map(void_to_zero, parsed_tree[4:])))
                    case_coordinates["Total"] = void_to_zero(parsed_tree[3])

        case_space_vectors.append(case_coordinates)
        basis |= set(attributes)

    case_space_matrix = np.array([[0. for _ in case_space_vectors] for _ in basis])
    case_vector = np.array([0. for _ in basis])

    for row, b in enumerate(basis):
        case_vector[row] = case_dict.get(b, 0.) / case_dict["Total"] if case_dict["Total"] != 0. else 0.
        for column, vector in enumerate(case_space_vectors):
            case_space_matrix[row, column] = vector.get(b, 0.) / vector["Total"] if vector["Total"] != 0. else 0.

    return case_space_matrix, case_vector


def case_space_case_angle_csv(treebank, treebank_case):
    case_space_matrix, case_vector = enhanced_get_matrix_csv(treebank, treebank_case)
    dimension = scipy.linalg.orth(case_space_matrix)
    if dimension.shape[1]:
        a = angle(case_vector, dimension)
    else:
        a = np.nan
    return treebank, treebank_case[0] + f"_Case={treebank_case[1]}", a


def compute_angles_csv():
    all_banks = get_all_banks()
    basis = overall_basis_csv()
    angles = []
    for corpus in tqdm(all_banks, colour="#7d1dd3", leave=True, desc="Computing Angles"):
        case_space = []
        for csv in filter(lambda t: t[-4:] == ".csv", os.listdir("RelDep_Matches")):
            with open(f"RelDep_Matches/{csv}", "r") as csv_file:
                attributes = next(csv_file).rstrip().split(",")[4:]
                for tree in filter(lambda t: t[0] == corpus, csv_file):
                    case = tree.rstrip().split(",")
                    coordinates = dict(zip(attributes, map(void_to_zero, case[4:])))
                    coordinates["Total"] = void_to_zero(case[3])
            case_space.append(coordinates)

        matrix = np.array([[0. for _ in case_space] for _ in basis])
        for row, b in enumerate(basis):
            for column, case_vec in enumerate(case_space):
                matrix[row, column] = case_vec.get(b, 0.) / case_vec["Total"] if case_vec["Total"] != 0. else 0.

        def get_angle(treebank, case_space_matrix, vector_filename):
            with open(f"RelDep_Matches/RelDep_matching_Case={vector_filename[1]}.csv", "r") as csvfile:
                vec_attr = next(csvfile).rstrip().split(",")[4:]
                for treebank_vector in filter(lambda t: t[0] == vector_filename[0], csvfile):
                    vec = treebank_vector.rstrip().split(",")
                    case_dict = dict(zip(vec_attr, map(void_to_zero, vec[4:])))
                    case_dict["Total"] = void_to_zero(vec[3])

            vector = [0. for _ in basis]
            for coordinate, base_vector in enumerate(basis):
                vector[coordinate] = case_dict.get(base_vector, 0.) / case_vec["Total"] if case_vec[
                                                                                               "Total"] != 0. else 0.

            dimension = scipy.linalg.orth(case_space_matrix)
            if dimension.shape[1]:
                a = angle(vector, dimension)
            else:
                a = np.nan
            return treebank, vector_filename[0] + f"_Case={vector_filename[1]}", a

        with tqdm_joblib(
                tqdm(
                    desc=f"Angles for {corpus}", leave=False, colour="#ffe500", position=1,
                    total=len(all_banks) * len(gf)
                )
        ) as progress_bar:
            angles.append(
                dict(
                    joblib.Parallel(n_jobs=8, verbose=0)(
                        joblib.delayed(get_angle)(corpus, matrix, [all_banks[j], c]) for j in range(len(all_banks))
                        for c in gf
                    )
                )
            )
        pandas.DataFrame(angles).to_csv(f"Proximities/Vector_Angle_Proximity.csv", index=False)


def tabulize_csv(grammar_feature):
    try:
        with open(f"Proximities/Proximity_Stats_for_{grammar_feature[0]}={grammar_feature[1]}"):
            pass
    except FileNotFoundError:
        print(f"Tabulizing for {grammar_feature[0]}={grammar_feature[1]}")
        treebanks = []
        with open(f"RelDep_Matches/RelDep_matching_{grammar_feature[0]}={grammar_feature[1]}.csv") as f:
            header = next(f).split(",")[4:]
            lines = f.readlines()
            mat = np.array([[0. for _ in lines] for _ in header])
            for row, line in enumerate(lines):
                reldep_frequencies = line.rstrip().split(",")
                treebanks.append(reldep_frequencies[0])
                coordinates = np.array(list(map(void_to_zero, reldep_frequencies[4:])))
                total = void_to_zero(reldep_frequencies[3])
                if total != 0.:
                    mat[:, row] = coordinates / total

        for i in range(len(mat[0])):
            euclidean_norm = npl.norm(mat[:, i])
            if euclidean_norm != 0:
                mat[:, i] /= euclidean_norm

        dot_mat = np.matmul(np.transpose(mat), mat)
        pandas.DataFrame(data=dot_mat, columns=treebanks).to_csv(
            f"Proximities/Proximity_Stats_for_{grammar_feature[0]}={grammar_feature[1]}.csv"
            )


gf = ["Nom", "Acc", "Dat", "Gen", "Voc", "Loc", "Abl", "Abs", "Erg"]

if __name__ == "__main__":
    for g in gf:
        tabulize_csv(("Case", g))
    # compute_distances_xl()
    # compute_vector_case_space_angles_xl()
    # print(case_space_case_angle_csv("hu_szeged-ud-dev", ["sah_yktdt-ud-test", "Acc"]))
